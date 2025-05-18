import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog
import time
import random
import psutil
import os

# Không có thông tin
from uninformed_search.bfs import bfs
from uninformed_search.dfs import dfs
from uninformed_search.UCS import ucs
from uninformed_search.IDS import ids

# Có thông tin
from Informed_search.greedy import greedy
from Informed_search.astar import a_star
from Informed_search.idastar import ida_star

#local search
from Local_search.Best_hill_climbing import steepest_ascent_hill_climbing
from Local_search.Simulated_Annealing import simulated_annealing
from Local_search.Beam_Search import beam_search
from Local_search.Genetic import genetic_algorithm

from Complex_enviroments.AndOr_search import and_or_search
from backtracking_gui import show_backtracking_gui
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
# Create initial and goal states
# 0 represents the empty tile in the puzzle
INITIAL_STATE = [[1, 2, 3], [4, 5, 6], [0, 7, 8]]
GOAL_STATE = [[1, 2, 3], [4, 5, 6], [7, 8, 0]]

class PuzzleSolverVisualizer(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.master.title("8-Puzzle Solver")
        self.master.geometry("1300x700")  # Slightly taller for menu
        self.master.resizable(False, False)
        self.algorithm = tk.StringVar()
        self.speed = tk.DoubleVar()
        self.speed.set(0.5)
        self.grid(row=0, column=0)
        self.current_state = [row[:] for row in INITIAL_STATE]
        self.initial_state = [row[:] for row in INITIAL_STATE]  # Store initial state
        self.create_menu()  # Add menu creation
        self.create_widgets()
        self.steps = 0
        self.is_solving = False
        self.should_stop = False
        self.current_path = None
        self.current_state_index = 0
        self.master.grid_columnconfigure(0, weight=0)
        self.master.grid_columnconfigure(1, weight=0)
        self.master.grid_columnconfigure(2, weight=0)

    def create_menu(self):
        # Create style for menu
        style = ttk.Style()
        style.configure('Menu.TFrame', background='#e1e1e1')
        
        # Create menu frame
        menu_frame = ttk.Frame(self.master, style='Menu.TFrame')
        menu_frame.grid(row=0, column=0, sticky="ew", columnspan=3)
        
        # Create and configure menubar with custom colors
        menubar = tk.Menu(self.master, bg='#f0f0f0', activebackground='#0078d7', activeforeground='white', font=('Arial', 10))
        self.master.config(menu=menubar)

        # File Menu with custom colors and styling
        file_menu = tk.Menu(menubar, tearoff=0, bg='#f0f0f0', fg='#000000',
                           activebackground='#0078d7', activeforeground='white',
                           font=('Arial', 10))
        menubar.add_cascade(label="File", menu=file_menu, font=('Arial', 10, 'bold'))
        file_menu.add_command(label="New", command=self.reset_puzzle)
        file_menu.add_command(label="Save State", command=self.save_state)
        file_menu.add_command(label="Load State", command=self.load_state)
        file_menu.add_command(label="Export Solution", command=self.export_solution)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.master.quit)

        # Help Menu with custom colors and styling
        help_menu = tk.Menu(menubar, tearoff=0, bg='#f0f0f0', fg='#000000',
                           activebackground='#0078d7', activeforeground='white',
                           font=('Arial', 10))
        menubar.add_cascade(label="Help", menu=help_menu, font=('Arial', 10, 'bold'))
        help_menu.add_command(label="Instructions", command=self.show_instructions)
        help_menu.add_command(label="About", command=self.show_about)

    def create_widgets(self):
        # Create a main frame to hold everything
        main_frame = ttk.Frame(self)
        main_frame.grid(row=1, column=0, sticky="nsew")  # Changed from row=0 to row=1 to account for menu
        
        # Thêm viền cho left_frame và right_frame
        left_frame = ttk.Frame(main_frame, relief="groove", padding=10)
        left_frame.grid(row=0, column=0, padx=10, pady=5, sticky="nsew")

        right_frame = ttk.Frame(main_frame, relief="groove", padding=10)
        right_frame.grid(row=0, column=1, padx=10, pady=5, sticky="nsew")

        # Add algorithm selection to left frame with proper padding
        algorithm_frame = ttk.Frame(left_frame)
        algorithm_frame.grid(row=0, column=0, sticky="w", pady=5)
        
        algorithm_label = ttk.Label(algorithm_frame, text="Choose algorithm:")
        algorithm_label.grid(row=0, column=0, sticky="w", padx=(0,10))
        
        # Create a smaller OptionMenu with fixed width
        algorithm_menu = ttk.OptionMenu(algorithm_frame, self.algorithm, "BFS", 
            "BFS", "DFS", "UCS", "IDS", "Greedy", "A*", "ID_A*", "Best Hill Climbing", 
            "Simulated Annealing", "Beam Search", "Genetic", "And-Or search", "Backtracking","Bactracking + AC3", "Q Learning")
        algorithm_menu.grid(row=0, column=1, sticky="w")
        algorithm_menu.configure(width=20)  # Set fixed width for the menu

        # Add speed control to left frame
        speed_label = ttk.Label(left_frame, text="Adjust speed:")
        speed_label.grid(row=2, column=0, columnspan=3, sticky="ew")
        speed_scale = ttk.Scale(left_frame, variable=self.speed, from_=0.1, to=1.0, orient=tk.HORIZONTAL, length=200)
        speed_scale.grid(row=3, column=0, columnspan=3)

        # Create a frame for buttons in left frame
        button_frame = ttk.Frame(left_frame)
        button_frame.grid(row=4, column=0, columnspan=3, sticky="ew", pady=5)
        
        # Create buttons in the button frame
        self.start_button = ttk.Button(button_frame, text="Start", command=self.start_solver)
        self.start_button.grid(row=0, column=0, padx=5)

        shuffle_button = ttk.Button(button_frame, text="Shuffle", command=self.shuffle_state)
        shuffle_button.grid(row=0, column=1, padx=5)

        self.stop_button = ttk.Button(button_frame, text="Stop", command=self.stop_solver, state='disabled')
        self.stop_button.grid(row=0, column=2, padx=5)

        self.continue_button = ttk.Button(button_frame, text="Continue", command=self.continue_solver, state='disabled')
        self.continue_button.grid(row=0, column=3, padx=5)

        # Add Reset and Clear Metrics buttons to button frame
        self.reset_button = ttk.Button(button_frame, text="Reset", command=self.reset_puzzle)
        self.reset_button.grid(row=0, column=4, padx=5)

        # self.clear_metrics_button = ttk.Button(button_frame, text="Clear Metrics", command=self.clear_metrics)
        # self.clear_metrics_button.grid(row=0, column=5, padx=5)

        # Add manual control buttons
        self.next_step_button = ttk.Button(button_frame, text="Next Step", command=self.next_step, state='disabled')
        self.next_step_button.grid(row=0, column=5, padx=5)

        self.back_step_button = ttk.Button(button_frame, text="Back Step", command=self.back_step, state='disabled')
        self.back_step_button.grid(row=0, column=6, padx=5)

        # Add canvas to left frame
        self.canvas = tk.Canvas(left_frame, width=300, height=300, borderwidth=0, relief=tk.FLAT)
        self.canvas.grid(row=5, column=0, columnspan=3, pady=10)

        # Initialize puzzle board
        box_width = 100
        box_height = 100
        for i in range(3):
            for j in range(3):
                x1 = j * box_width
                y1 = i * box_height
                x2 = x1 + box_width
                y2 = y1 + box_height
                number = INITIAL_STATE[i][j]
                if number != 0:
                    self.canvas.create_rectangle(x1, y1, x2, y2, fill='white', outline='black')
                    self.canvas.create_text(x1 + box_width//2, y1 + box_height//2, text=str(number))
                else:
                    self.canvas.create_rectangle(x1, y1, x2, y2, fill='gray', outline='black')

        # Add metrics frame to left frame
        metrics_frame = ttk.LabelFrame(left_frame, text="Performance Metrics")
        metrics_frame.grid(row=6, column=0, columnspan=3, sticky="ew", pady=5)

        self.steps_label = ttk.Label(metrics_frame, text="Steps: -")
        self.steps_label.grid(row=0, column=0, padx=10, pady=2, sticky="w")

        self.time_label = ttk.Label(metrics_frame, text="Time: -")
        self.time_label.grid(row=1, column=0, padx=10, pady=2, sticky="w")

        self.memory_label = ttk.Label(metrics_frame, text="Memory: -")
        self.memory_label.grid(row=2, column=0, padx=10, pady=2, sticky="w")

        # Add status bar to left frame
        self.status_bar = ttk.Label(left_frame, text="Ready", relief=tk.SUNKEN, anchor="w")
        self.status_bar.grid(row=7, column=0, columnspan=3, sticky="ew", pady=5)

        # Add state entry section to right frame
        enter_state_button = ttk.Button(right_frame, text="Enter State", command=self.enter_state)
        enter_state_button.grid(row=0, column=0, pady=5)

        # Add state display section to right frame
        state_frame = ttk.LabelFrame(right_frame, text="States")
        state_frame.grid(row=1, column=0, sticky="ew", pady=5)

        # Add start state canvas
        start_state_label = ttk.Label(state_frame, text="Start State:")
        start_state_label.grid(row=0, column=0, pady=5)
        self.start_canvas = tk.Canvas(state_frame, width=150, height=150, borderwidth=2, relief=tk.GROOVE)
        self.start_canvas.grid(row=1, column=0, padx=10, pady=5)
        self.display_start_state()

        # Add goal state canvas NGANG với start state
        goal_state_label = ttk.Label(state_frame, text="Goal State:")
        goal_state_label.grid(row=0, column=1, pady=5)
        self.goal_canvas = tk.Canvas(state_frame, width=150, height=150, borderwidth=2, relief=tk.GROOVE)
        self.goal_canvas.grid(row=1, column=1, padx=10, pady=5)
        self.display_goal_state()

        # Add Algorithm Statistic table (Treeview) to right_frame
        statistic_frame = ttk.LabelFrame(right_frame, text="Algorithm Statistic")
        statistic_frame.grid(row=2, column=0, sticky="ew", pady=10)
        stat_columns = ("Algorithm", "Time (s)", "Steps", "Cost", "Space")
        self.statistic_table = ttk.Treeview(statistic_frame, columns=stat_columns, show="headings", height=13)
        for col in stat_columns:
            self.statistic_table.heading(col, text=col)
            self.statistic_table.column(col, width=120, anchor="center")
        self.statistic_table.pack(fill="x")

        # Configure grid weights
        main_frame.grid_columnconfigure(1, weight=1)
        left_frame.grid_columnconfigure(0, weight=1)
        right_frame.grid_columnconfigure(0, weight=1)

        # Add tooltips for algorithms
        algorithm_menu.bind('<Enter>', lambda e: self.show_algorithm_tooltip(e))
        algorithm_menu.bind('<Leave>', lambda e: self.hide_tooltip())

    def get_inversion_count(self, state):
        # Flatten the state and remove the blank tile (0)
        flat_state = [num for row in state for num in row if num != 0]
        inversion_count = 0
        
        # Count inversions
        for i in range(len(flat_state)):
            for j in range(i + 1, len(flat_state)):
                if flat_state[i] > flat_state[j]:
                    inversion_count += 1
        
        return inversion_count

    def is_solvable(self, state):
        # For 3x3 puzzle (odd dimension), the puzzle is solvable 
        # if number of inversions is even
        inversion_count = self.get_inversion_count(state)
        return inversion_count % 2 == 0

    def shuffle_state(self):
        # Keep shuffling until we get a solvable state
        while True:
            # Flatten the 2D array, shuffle it, and reshape back to 3x3
            numbers = [num for row in self.current_state for num in row]
            random.shuffle(numbers)
            new_state = [numbers[i:i+3] for i in range(0, 9, 3)]
            
            # Check if the shuffled state is solvable
            if self.is_solvable(new_state):
                self.current_state = new_state
                break
        
        # Update both main display and start state display
        self.display_state()
        self.display_start_state()

    def display_state(self):
        # Clear the canvas
        self.canvas.delete("all")
        
        # Define box dimensions
        box_width = 100
        box_height = 100
        
        # Draw the current state
        for i in range(3):
            for j in range(3):
                x1 = j * box_width
                y1 = i * box_height
                x2 = x1 + box_width
                y2 = y1 + box_height
                number = self.current_state[i][j]
                if number != 0:
                    self.canvas.create_rectangle(x1, y1, x2, y2, fill='white', outline='black', width=2)
                    self.canvas.create_text(x1 + box_width//2, y1 + box_height//2, text=str(number), font=('Arial', 24))
                else:
                    self.canvas.create_rectangle(x1, y1, x2, y2, fill='gray', outline='gray', width=2)
        
        # Update the canvas
        self.canvas.update()

    def stop_solver(self):
        self.should_stop = True
        self.stop_button.configure(state='disabled')
        self.continue_button.configure(state='normal')
        self.start_button.configure(state='normal')
        # Enable manual step controls when stopped
        if hasattr(self, 'current_path') and self.current_path:
            self.next_step_button.configure(state='normal')
            self.back_step_button.configure(state='normal')
        else:
            self.next_step_button.configure(state='disabled')
            self.back_step_button.configure(state='disabled')

    def continue_solver(self):
        self.should_stop = False
        self.stop_button.configure(state='normal')
        self.continue_button.configure(state='disabled')
        self.start_button.configure(state='disabled')
        # Disable manual step controls when continuing
        self.next_step_button.configure(state='disabled')
        self.back_step_button.configure(state='disabled')
        if self.current_path:
            self.display_path(self.current_path, start_from_index=self.current_state_index)

    def start_solver(self):
        if self.algorithm.get() == "Backtracking":
            show_backtracking_gui()
            return
            
        if self.is_solving:
            return

        self.status_bar.config(text="Solving...")
        algorithm = self.algorithm.get()
        self.master.title(f"8-Puzzle Solver - Running {algorithm}")
        # Check if current state is solvable
        # if not self.is_solvable(self.current_state):
        #     messagebox.showinfo("Unsolvable State", "Current state is not solvable. Number of inversions is odd.")
        #     return

        # Get the chosen algorithm and speed
        algorithm = self.algorithm.get()
        
        # Reset state
        self.should_stop = False
        self.current_state_index = 0
        self.stop_button.configure(state='normal')
        self.continue_button.configure(state='disabled')
        self.start_button.configure(state='disabled')

        # Get initial memory usage
        process = psutil.Process(os.getpid())
        initial_memory = process.memory_info().rss
        
        # Start timing
        start_time = time.time()

        # Solve the puzzle using the chosen algorithm
        if algorithm == "BFS":
            path,cost,space = bfs(self.current_state, GOAL_STATE)
        elif algorithm == "DFS":
            path,cost,space = dfs(self.current_state, GOAL_STATE)
        elif algorithm == "Greedy":
            path,cost,space = greedy(self.current_state, GOAL_STATE)
        elif algorithm == "A*":
            path,cost,space = a_star(self.current_state, GOAL_STATE)
        elif algorithm == "ID_A*":
            path,cost,space = ida_star(self.current_state,GOAL_STATE)
        elif algorithm == "Best Hill Climbing":
            path,cost,space = steepest_ascent_hill_climbing(self.current_state, GOAL_STATE)
        elif algorithm == "Simulated Annealing":
            path,cost,space = simulated_annealing(self.current_state, GOAL_STATE)
        elif algorithm == "Beam Search":
            path,cost,space = beam_search(self.current_state, GOAL_STATE)
        elif algorithm == "UCS":
            path,cost,space = ucs(self.current_state, GOAL_STATE)
        elif algorithm == "And-Or search":
            path,cost,space = and_or_search(self.current_state, GOAL_STATE)
        elif algorithm == "IDS":
            path,cost,space = ids(self.current_state, GOAL_STATE)  
        elif algorithm == "Genetic":
            path,cost,space = genetic_algorithm(self.current_state,GOAL_STATE)
        # Calculate metrics
        end_time = time.time()
        final_memory = process.memory_info().rss
        execution_time = end_time - start_time
        memory_used = (final_memory - initial_memory) / (1024 * 1024)  # Convert to MB

        # Display the path on the canvas
        if path is not None:
            self.steps = len(path)
            self.steps_label.config(text=f"Steps: {self.steps}")
            self.time_label.config(text=f"Time: {execution_time:.2f}s")
            self.memory_label.config(text=f"Memory: {abs(memory_used):.2f}MB")
            self.status_bar.config(text="Solution found!")
            self.current_path = path
            self.display_path(path)
            # Thêm/cập nhật vào bảng thống kê (cost và space để mặc định là '-')
            self.update_algorithm_statistic(algorithm, len(path), execution_time,cost,space)
        else:
            self.status_bar.config(text="No solution found!")
            messagebox.showinfo("Unsolvable Initial State", "Unsolvable Initial State, Solution cannot be found")
            self.stop_button.configure(state='disabled')
            self.start_button.configure(state='normal')

    def display_path(self, path, start_from_index=0):
        # Define box dimensions
        box_width = 100
        box_height = 100
    
        # Copy the initial state
        state = [row[:] for row in self.current_state]
        
        # Apply moves up to the start_from_index
        for i in range(start_from_index):
            action = path[i]
            self.apply_move(state, action)
    
        # Loop through remaining actions in the path
        for i in range(start_from_index, len(path)):
            if self.should_stop:
                self.current_state_index = i
                return
  
            action = path[i]
            self.apply_move(state, action)
            self.display_state_with_state(state)
    
            # Sleep to slow down the animation based on the selected speed
            time.sleep(1 - self.speed.get())
        
        # Animation completed
        self.stop_button.configure(state='disabled')
        self.continue_button.configure(state='disabled')
        self.start_button.configure(state='normal')
        self.master.title("8-Puzzle Solver")  # Reset title when done

    def apply_move(self, state, action):
        if action == "Up":
            row, col = self.find_blank(state)
            if row > 0:
                state[row][col], state[row-1][col] = state[row-1][col], state[row][col]
        elif action == "Down":
            row, col = self.find_blank(state)
            if row < 2:
                state[row][col], state[row+1][col] = state[row+1][col], state[row][col]
        elif action == "Left":
            row, col = self.find_blank(state)
            if col > 0:
                state[row][col], state[row][col-1] = state[row][col-1], state[row][col]
        elif action == "Right":
            row, col = self.find_blank(state)
            if col < 2:
                state[row][col], state[row][col+1] = state[row][col+1], state[row][col]

    def display_state_with_state(self, state):
        # Clear the canvas
        self.canvas.delete("all")
        
        # Define box dimensions
        box_width = 100
        box_height = 100
        
        # Draw the current state
        for i in range(3):
            for j in range(3):
                x1 = j * box_width
                y1 = i * box_height
                x2 = x1 + box_width
                y2 = y1 + box_height
                number = state[i][j]
                if number != 0:
                    self.canvas.create_rectangle(x1, y1, x2, y2, fill='white', outline='black')
                    self.canvas.create_text(x1 + box_width//2, y1 + box_height//2, text=str(number))
                else:
                    self.canvas.create_rectangle(x1, y1, x2, y2, fill='gray', outline='black')
        
        # Update the canvas
        self.canvas.update()

    def find_blank(self, state):
        # Find the row and column of the blank tile
        for row in range(3):
            for col in range(3):
                if state[row][col] == 0:
                    return row, col

    def enter_state(self):
        def submit_state():
            try:
                # Lấy dữ liệu từ các ô nhập
                new_state = []
                for i in range(3):
                    row = []
                    for j in range(3):
                        value = int(entries[i][j].get())
                        if value < 0 or value > 8:
                            raise ValueError("Values must be between 0 and 8.")
                        row.append(value)
                    new_state.append(row)

                # Kiểm tra tính hợp lệ của trạng thái
                if sorted([num for row in new_state for num in row]) != list(range(9)):
                    raise ValueError("State must contain all numbers from 0 to 8.")

                # Cập nhật trạng thái hiện tại
                if self.is_solvable(new_state):
                    self.current_state = new_state
                    self.display_state()
                    self.display_start_state()  # Cập nhật trạng thái bắt đầu
                    messagebox.showinfo("Success", "Initial state updated successfully!")
                    input_window.destroy()
                else:
                    messagebox.showerror("Error", "The entered state is not solvable.")
            except ValueError as e:
                messagebox.showerror("Error", f"Invalid input: {e}")

        # Tạo cửa sổ nhập liệu
        input_window = tk.Toplevel(self.master)
        input_window.title("Enter Initial State")
        input_window.geometry("300x250")  # Điều chỉnh kích thước cửa sổ
        
        # Tạo frame chứa các ô nhập liệu
        input_frame = ttk.Frame(input_window, padding="20")
        input_frame.grid(row=0, column=0, sticky="nsew")
        
        # Thêm label hướng dẫn
        instruction_label = ttk.Label(input_frame, text="Enter numbers (0-8):", font=('Arial', 10))
        instruction_label.grid(row=0, column=0, columnspan=3, pady=(0, 10))

        # Tạo các ô nhập liệu
        entries = []
        for i in range(3):
            row_entries = []
            for j in range(3):
                entry = ttk.Entry(input_frame, width=3, justify="center", font=('Arial', 14))
                entry.grid(row=i+1, column=j, padx=5, pady=5, ipady=5)  # Thêm ipady để ô cao hơn
                row_entries.append(entry)
            entries.append(row_entries)

        # Thêm nút Submit với style mới
        submit_button = ttk.Button(input_frame, text="Submit", command=submit_state)
        submit_button.grid(row=4, column=0, columnspan=3, pady=20)

        # Cấu hình grid
        input_window.grid_columnconfigure(0, weight=1)
        input_window.grid_rowconfigure(0, weight=1)
        for i in range(3):
            input_frame.grid_columnconfigure(i, weight=1)

        # Đặt cửa sổ ở giữa cửa sổ chính
        input_window.transient(self.master)
        input_window.grab_set()
        self.master.eval(f'tk::PlaceWindow {input_window} center')

    def display_start_state(self):
        # Clear the canvas
        self.start_canvas.delete("all")

        # Define box dimensions
        box_width = 50
        box_height = 50

        # Draw the current start state
        for i in range(3):
            for j in range(3):
                x1 = j * box_width
                y1 = i * box_height
                x2 = x1 + box_width
                y2 = y1 + box_height
                number = self.current_state[i][j]
                if number != 0:
                    self.start_canvas.create_rectangle(x1, y1, x2, y2, fill='white', outline='black')
                    self.start_canvas.create_text(x1 + box_width//2, y1 + box_height//2, text=str(number))
                else:
                    self.start_canvas.create_rectangle(x1, y1, x2, y2, fill='gray', outline='black')

        # Update the canvas
        self.start_canvas.update()

    def display_goal_state(self):
        # Clear the canvas
        self.goal_canvas.delete("all")

        # Define box dimensions
        box_width = 50
        box_height = 50

        # Draw the goal state
        for i in range(3):
            for j in range(3):
                x1 = j * box_width
                y1 = i * box_height
                x2 = x1 + box_width
                y2 = y1 + box_height
                number = GOAL_STATE[i][j]
                if number != 0:
                    self.goal_canvas.create_rectangle(x1, y1, x2, y2, fill='white', outline='black')
                    self.goal_canvas.create_text(x1 + box_width//2, y1 + box_height//2, text=str(number))
                else:
                    self.goal_canvas.create_rectangle(x1, y1, x2, y2, fill='gray', outline='black')

        # Update the canvas
        self.goal_canvas.update()

    def reset_puzzle(self):
        """Reset the puzzle to its initial state"""
        self.current_state = [row[:] for row in self.initial_state]
        self.current_state_index = 0
        self.display_state()
        self.display_start_state()
        self.clear_metrics()
        self.status_bar.config(text="Ready")
        # Disable manual step controls when resetting
        self.next_step_button.configure(state='disabled')
        self.back_step_button.configure(state='disabled')

    def clear_metrics(self):
        """Clear all performance metrics"""
        self.steps = 0
        self.steps_label.config(text="Steps: -")
        self.time_label.config(text="Time: -")
        self.memory_label.config(text="Memory: -")

    def save_state(self):
        """Save current state to a file"""
        try:
            with open("puzzle_state.txt", "w") as f:
                for row in self.current_state:
                    f.write(",".join(map(str, row)) + "\n")
            messagebox.showinfo("Success", "State saved successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Could not save state: {e}")

    def load_state(self):
        """Load state from a file"""
        try:
            with open("puzzle_state.txt", "r") as f:
                new_state = []
                for line in f:
                    row = list(map(int, line.strip().split(",")))
                    new_state.append(row)
                if self.is_solvable(new_state):
                    self.current_state = new_state
                    self.display_state()
                    self.display_start_state()
                else:
                    messagebox.showerror("Error", "Loaded state is not solvable")
        except Exception as e:
            messagebox.showerror("Error", f"Could not load state: {e}")

    def show_instructions(self):
        """Show instructions dialog"""
        instructions = """
8-Puzzle Solver Instructions:

1. Choose an algorithm from the dropdown menu
2. Adjust the animation speed using the slider
3. Click 'Shuffle' to randomize the puzzle
4. Click 'Start' to begin solving
5. Use 'Stop' and 'Continue' to control the animation
6. Click 'Reset' to return to initial state

You can also:
- Enter your own initial state
- Save and load puzzle states
- Compare different algorithms
"""
        messagebox.showinfo("Instructions", instructions)

    def show_about(self):
        """Show about dialog"""
        about_text = """
8-Puzzle Solver
Version 1.0

A visual tool for solving the 8-puzzle problem using various algorithms.

Supported algorithms:
- BFS (Breadth-First Search)
- DFS (Depth-First Search)
- A* (A-Star Search)
- And more...
"""
        messagebox.showinfo("About", about_text)

    def show_algorithm_tooltip(self, event):
        """Show tooltip explaining the current algorithm"""
        algorithm = self.algorithm.get()
        tooltips = {
            "BFS": "Breadth-First Search: Explores all nodes at present depth before moving to nodes at next depth",
            "DFS": "Depth-First Search: Explores as far as possible along each branch before backtracking",
            "A*": "A* Search: Uses heuristics to find the optimal path",
            "Greedy": "Greedy Search: Always chooses the path that looks best at the moment",
            "ID_A*": "Iterative Deepening A*: Combines depth-first search's space-efficiency with breadth-first search's completeness",
            "Best Hill Climbing": "Best Hill Climbing: Always moves in the direction of increasing value",
            "Simulated Annealing": "Simulated Annealing: Probabilistically accepts worse solutions to escape local optima",
            "Beam Search": "Beam Search: Breadth-first search with limited memory",
            "UCS": "Uniform Cost Search: Finds the path with the lowest cost",
            "And-Or tree": "And-Or Tree Search: Handles nondeterministic actions"
        }
        
        tooltip = tooltips.get(algorithm, "Select an algorithm")
        
        # Create tooltip window
        self.tooltip = tk.Toplevel()
        self.tooltip.wm_overrideredirect(True)
        
        # Create frame with border
        frame = ttk.Frame(self.tooltip, relief="solid", borderwidth=1)
        frame.pack()
        
        # Create label with smaller width and font
        label = ttk.Label(frame, text=tooltip, justify='left',
                         background="#ffffe0", wraplength=250,
                         font=('Arial', 9))
        label.pack(padx=5, pady=3)
        
        # Position tooltip relative to algorithm menu
        x = event.widget.winfo_rootx()
        y = event.widget.winfo_rooty() + event.widget.winfo_height()
        self.tooltip.wm_geometry(f"+{x}+{y}")

    def hide_tooltip(self):
        """Hide the algorithm tooltip"""
        if hasattr(self, 'tooltip'):
            self.tooltip.destroy()

    def export_solution(self):
        """Export solution steps to a file"""
        if not hasattr(self, 'current_path') or not self.current_path:
            messagebox.showinfo("No Solution", "No solution to export. Please solve the puzzle first.")
            return
        
        try:
            # Tạo tên file mặc định dựa trên thuật toán
            algorithm = self.algorithm.get()
            default_filename = f"8Puzzle_Solution_{algorithm}.xlsx"
            
            # Hiển thị hộp thoại để chọn vị trí lưu file
            file_path = tk.filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Save Solution Steps",
                initialfile=default_filename
            )
            
            if not file_path:  # Nếu người dùng bấm Cancel
                return

            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Solution Steps"

            # Set column widths
            for col in range(1, 10):
                ws.column_dimensions[get_column_letter(col)].width = 5

            # Write headers
            ws['A1'] = "8-Puzzle Solution Steps"
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:I1')

            # Write initial state
            ws['A3'] = "Initial State:"
            ws['A3'].font = Font(bold=True)
            for i in range(3):
                for j in range(3):
                    cell = ws.cell(row=i+4, column=j+1)
                    cell.value = self.current_state[i][j]
                    cell.alignment = Alignment(horizontal='center')
                    if self.current_state[i][j] == 0:
                        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            # Write goal state
            ws['A8'] = "Goal State:"
            ws['A8'].font = Font(bold=True)
            for i in range(3):
                for j in range(3):
                    cell = ws.cell(row=i+9, column=j+1)
                    cell.value = GOAL_STATE[i][j]
                    cell.alignment = Alignment(horizontal='center')
                    if GOAL_STATE[i][j] == 0:
                        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            # Write solution steps
            ws['A13'] = "Solution Steps:"
            ws['A13'].font = Font(bold=True)
            
            # Create a copy of initial state to track changes
            current_state = [row[:] for row in self.current_state]
            
            # Write each step
            for step_idx, action in enumerate(self.current_path, 1):
                # Write step number and action
                ws.cell(row=step_idx*5+13, column=1, value=f"Step {step_idx}:")
                ws.cell(row=step_idx*5+13, column=2, value=f"Move blank tile {action}")
                
                # Apply the move
                self.apply_move(current_state, action)
                
                # Write the state after the move as a 3x3 matrix
                for i in range(3):
                    for j in range(3):
                        cell = ws.cell(row=step_idx*5+13+i+1, column=j+1)
                        cell.value = current_state[i][j]
                        cell.alignment = Alignment(horizontal='center')
                        if current_state[i][j] == 0:
                            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

            # Write total steps
            total_steps_row = len(self.current_path) * 5 + 17
            ws.cell(row=total_steps_row, column=1, value="Total Steps:")
            ws.cell(row=total_steps_row, column=2, value=len(self.current_path))
            ws.cell(row=total_steps_row, column=1).font = Font(bold=True)

            # Save the workbook
            wb.save(file_path)
            messagebox.showinfo("Success", f"Solution steps exported to '{file_path}'!")
        except Exception as e:
            messagebox.showerror("Error", f"Could not export solution: {e}")

    def next_step(self):
        """Move one step forward in the solution"""
        if not hasattr(self, 'current_path') or not self.current_path:
            return
        
        if self.current_state_index < len(self.current_path):
            # Get the next action
            action = self.current_path[self.current_state_index]
            
            # Apply the move
            self.apply_move(self.current_state, action)
            self.display_state()
            
            # Update index
            self.current_state_index += 1
            
            # Update button states
            self.back_step_button.configure(state='normal')
            if self.current_state_index >= len(self.current_path):
                self.next_step_button.configure(state='disabled')

    def back_step(self):
        """Move one step backward in the solution"""
        if not hasattr(self, 'current_path') or not self.current_path:
            return
        
        if self.current_state_index > 0:
            # Giảm index trước
            self.current_state_index -= 1
            
            # Lấy hành động tại vị trí hiện tại và tạo hành động ngược lại
            current_action = self.current_path[self.current_state_index]
            reverse_action = {
                "Up": "Down",
                "Down": "Up",
                "Left": "Right",
                "Right": "Left"
            }[current_action]
            
            # Áp dụng hành động ngược lại
            self.apply_move(self.current_state, reverse_action)
            self.display_state()
            
            # Cập nhật trạng thái các nút
            self.next_step_button.configure(state='normal')
            if self.current_state_index == 0:
                self.back_step_button.configure(state='disabled')

    def update_algorithm_statistic(self, algorithm, steps, execution_time, cost="-", space="-"):
        """Cập nhật hoặc thêm một dòng vào bảng thống kê thuật toán (mỗi thuật toán chỉ 1 dòng)"""
        # Tìm xem đã có dòng cho thuật toán này chưa
        for item in self.statistic_table.get_children():
            if self.statistic_table.item(item, 'values')[0] == algorithm:
                self.statistic_table.item(item, values=(algorithm, f"{execution_time:.2f}", steps, cost, space))
                return
        # Nếu chưa có thì thêm mới
        self.statistic_table.insert("", "end", values=(algorithm, f"{execution_time:.2f}", steps, cost, space))

def main():
   root = tk.Tk()
   root.title("8 Puzzle Visualizer")
   app = PuzzleSolverVisualizer(master=root)
   app.mainloop()   

if __name__ == "__main__":
    main()